"""
Check the Excel file to see all rows in the Roles sheet
"""

import openpyxl

EXCEL_PATH = r'C:\Users\jomon\Documents\MyDocuments\CourseWork\Thesis\SE-QPT.xlsx'

try:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = wb['Roles']

    print(f"\n=== Roles Sheet Analysis ===\n")
    print(f"Total rows (including empty): {sheet.max_row}\n")

    # Read all rows including potential empty ones
    for row_idx in range(1, sheet.max_row + 5):  # Read a few extra rows just in case
        col1 = sheet.cell(row_idx, 1).value
        col2 = sheet.cell(row_idx, 2).value

        if col1 or col2:  # If any cell has content
            print(f"Row {row_idx}:")
            print(f"  Column A: {col1}")
            print(f"  Column B: {col2[:100] if col2 and len(str(col2)) > 100 else col2}")
            print()

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
