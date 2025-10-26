"""
Compare role-competency matrix data from three sources:
1. Excel source of truth (Qualifizierungsmodule_Qualifizierungspläne_v4_enUS.xlsx)
2. Derik's original database dump
3. Current database

This will help identify discrepancies causing incorrect role suggestions.
"""

import os
import sys
import openpyxl
import psycopg2
from collections import defaultdict

# Database connection
DATABASE_URL = os.getenv('DATABASE_URL', 'postgresql://ma0349:MA0349_2025@localhost:5432/competency_assessment')

# File paths
EXCEL_PATH = '../../data/source/excel/Qualifizierungsmodule_Qualifizierungspläne_v4_enUS.xlsx'
DERIK_SQL = '../../../sesurveyapp-main/postgres-init/init.sql'

def read_excel_matrix():
    """
    Read the Role-Competences-Matrix sheet from the Excel source of truth.
    Returns: dict mapping (role_name, competency_name) -> value

    Excel structure:
    - Row 2: Role names starting from column 4
    - Row 3+: Competency data
    - Column 2: Competency names
    - Columns 4+: Values for each role
    """
    print("\n=== Reading Excel Source of Truth ===")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    sheet = wb['Role-Competences-Matrix']
    print(f"Found sheet: Role-Competences-Matrix")

    excel_data = {}

    # Read role names from row 2, starting at column 4
    roles = []
    for col_idx in range(4, sheet.max_column + 1):
        cell = sheet.cell(2, col_idx)
        if cell.value:
            role_name = str(cell.value).strip()
            roles.append((col_idx, role_name))

    print(f"Found {len(roles)} roles in Excel:")
    for _, role in roles:
        print(f"  - {role}")

    # Read competency rows starting from row 3
    competency_count = 0
    for row_idx in range(3, sheet.max_row + 1):
        # Column 2 has competency name
        competency_cell = sheet.cell(row_idx, 2)
        if not competency_cell.value:
            continue

        competency_name = str(competency_cell.value).strip()
        if not competency_name or competency_name == 'None':
            continue

        competency_count += 1

        # Read values for each role
        for col_idx, role_name in roles:
            value_cell = sheet.cell(row_idx, col_idx)
            value = value_cell.value

            # Convert to numeric, default to 0
            try:
                value = int(value) if value is not None else 0
            except (ValueError, TypeError):
                value = 0

            if value > 0:  # Only store non-zero values
                excel_data[(role_name, competency_name)] = value

    print(f"Found {competency_count} competencies in Excel")
    print(f"Total non-zero entries: {len(excel_data)}")

    return excel_data


def read_current_database():
    """
    Read role_competency_matrix from current database.
    Returns: dict mapping (role_name, competency_name) -> value
    """
    print("\n=== Reading Current Database ===")

    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    # Get all role-competency mappings with names
    query = """
        SELECT
            rc.role_cluster_name as role_name,
            c.competency_name,
            rcm.role_competency_value,
            rcm.organization_id
        FROM role_competency_matrix rcm
        JOIN role_cluster rc ON rcm.role_cluster_id = rc.id
        JOIN competency c ON rcm.competency_id = c.id
        WHERE rcm.role_competency_value > 0
        ORDER BY rc.role_cluster_name, c.competency_name
    """

    cur.execute(query)
    rows = cur.fetchall()

    db_data = {}
    for role_name, competency_name, value, org_id in rows:
        # Use org_id 11 or 16 as primary (these are the main organizations)
        if org_id in (11, 16):
            key = (role_name.strip(), competency_name.strip())
            db_data[key] = value

    print(f"Found {len(db_data)} non-zero entries in current database")

    cur.close()
    conn.close()

    return db_data


def analyze_specific_roles(excel_data, db_data):
    """
    Analyze specific roles that failed in testing.
    """
    print("\n=== Analyzing Failed Test Cases ===")

    # Roles that failed in tests
    problem_roles = [
        'Specialist Developer',  # Expected for Senior Software Developer
        'Project Manager',       # Incorrectly suggested
        'System Engineer',       # Expected for Systems Integration Engineer
        'Internal Support',      # Incorrectly suggested
        'Quality Engineer/Manager',  # Expected for QA Specialist (might be named differently)
        'Production Planner/Coordinator'  # Incorrectly suggested
    ]

    # Get all unique role names from both sources
    excel_roles = set(k[0] for k in excel_data.keys())
    db_roles = set(k[0] for k in db_data.keys())

    print(f"\nRoles in Excel: {sorted(excel_roles)}")
    print(f"\nRoles in Database: {sorted(db_roles)}")

    for role in problem_roles:
        print(f"\n--- Role: {role} ---")

        # Check if role exists in both sources
        in_excel = any(k[0] == role for k in excel_data.keys())
        in_db = any(k[0] == role for k in db_data.keys())

        print(f"In Excel: {in_excel}, In DB: {in_db}")

        if not in_excel and not in_db:
            # Try fuzzy match
            excel_matches = [r for r in excel_roles if role.lower() in r.lower() or r.lower() in role.lower()]
            db_matches = [r for r in db_roles if role.lower() in r.lower() or r.lower() in role.lower()]
            print(f"Possible Excel matches: {excel_matches}")
            print(f"Possible DB matches: {db_matches}")
            continue

        # Get top competencies from both sources
        excel_comps = [(k[1], v) for k, v in excel_data.items() if k[0] == role]
        db_comps = [(k[1], v) for k, v in db_data.items() if k[0] == role]

        excel_comps.sort(key=lambda x: x[1], reverse=True)
        db_comps.sort(key=lambda x: x[1], reverse=True)

        print(f"\nTop 10 competencies in Excel:")
        for comp, val in excel_comps[:10]:
            print(f"  {comp}: {val}")

        print(f"\nTop 10 competencies in Database:")
        for comp, val in db_comps[:10]:
            print(f"  {comp}: {val}")


def compare_sources(excel_data, db_data):
    """
    Compare Excel and database sources for discrepancies.
    """
    print("\n=== Comparing Sources ===")

    # Find keys in Excel but not in DB
    excel_only = set(excel_data.keys()) - set(db_data.keys())
    # Find keys in DB but not in Excel
    db_only = set(db_data.keys()) - set(excel_data.keys())
    # Find keys in both with different values
    common_keys = set(excel_data.keys()) & set(db_data.keys())

    different_values = []
    for key in common_keys:
        if excel_data[key] != db_data[key]:
            different_values.append((key, excel_data[key], db_data[key]))

    print(f"\nEntries only in Excel: {len(excel_only)}")
    if excel_only and len(excel_only) <= 20:
        for key in list(excel_only)[:20]:
            print(f"  {key[0]} -> {key[1]}: {excel_data[key]}")

    print(f"\nEntries only in Database: {len(db_only)}")
    if db_only and len(db_only) <= 20:
        for key in list(db_only)[:20]:
            print(f"  {key[0]} -> {key[1]}: {db_data[key]}")

    print(f"\nEntries with different values: {len(different_values)}")
    if different_values and len(different_values) <= 20:
        for key, excel_val, db_val in different_values[:20]:
            print(f"  {key[0]} -> {key[1]}: Excel={excel_val}, DB={db_val}")


if __name__ == '__main__':
    print("=" * 80)
    print("ROLE-COMPETENCY MATRIX SOURCE COMPARISON")
    print("=" * 80)

    # Read sources
    excel_data = read_excel_matrix()
    db_data = read_current_database()

    # Compare
    compare_sources(excel_data, db_data)

    # Analyze specific problem roles
    analyze_specific_roles(excel_data, db_data)

    print("\n" + "=" * 80)
    print("Analysis complete. Check output above for discrepancies.")
    print("=" * 80)
