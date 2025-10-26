"""
Inspect the Excel file structure to understand the Role-Competences-Matrix sheet format.
"""

import openpyxl

EXCEL_PATH = '../../data/source/excel/Qualifizierungsmodule_Qualifizierungspläne_v4_enUS.xlsx'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
sheet = wb['Role-Competences-Matrix']

print("=== Role-Competences-Matrix Sheet Structure ===\n")
print(f"Max row: {sheet.max_row}")
print(f"Max column: {sheet.max_column}")

print("\n--- First 5 rows, first 10 columns ---\n")

for row_idx in range(1, min(6, sheet.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(11, sheet.max_column + 1)):
        cell = sheet.cell(row_idx, col_idx)
        value = str(cell.value) if cell.value is not None else '<empty>'
        # Truncate long values
        if len(value) > 30:
            value = value[:27] + '...'
        row_data.append(value)

    print(f"Row {row_idx}: {' | '.join(row_data)}")

print("\n--- First column (role names?) ---\n")

for row_idx in range(1, min(15, sheet.max_row + 1)):
    cell = sheet.cell(row_idx, 1)
    print(f"Row {row_idx}: {cell.value}")

print("\n--- First row (competency names?) ---\n")

for col_idx in range(1, min(15, sheet.max_column + 1)):
    cell = sheet.cell(1, col_idx)
    print(f"Col {col_idx}: {cell.value}")
