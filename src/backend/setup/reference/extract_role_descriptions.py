"""
Extract role descriptions from Excel file and update database
This script reads the 14 standard SE roles and their descriptions from the Excel file
and updates the role_cluster table in the database.
"""

import openpyxl
import sys
import os

# Add the backend directory to the path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import create_app
from models import db, RoleCluster

# Path to the Excel file
EXCEL_PATH = r'C:\Users\jomon\Documents\MyDocuments\CourseWork\Thesis\SE-QPT.xlsx'

def extract_role_descriptions():
    """Extract role names and descriptions from the Excel file"""
    print("\n=== Extracting Role Descriptions from Excel ===\n")

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

        # Check if 'Roles' sheet exists
        if 'Roles' not in wb.sheetnames:
            print(f"ERROR: 'Roles' sheet not found in Excel file")
            print(f"Available sheets: {wb.sheetnames}")
            return None

        sheet = wb['Roles']
        print(f"Found 'Roles' sheet with {sheet.max_row} rows\n")

        roles_data = []

        # Start from row 1 (no header row - data starts from row 1)
        for row_idx in range(1, sheet.max_row + 1):
            role_name = sheet.cell(row_idx, 1).value  # Column A - Role name
            description = sheet.cell(row_idx, 2).value  # Column B - Description

            if role_name and description:
                roles_data.append({
                    'name': str(role_name).strip(),
                    'description': str(description).strip()
                })
                print(f"Row {row_idx}: {role_name}")

        print(f"\nExtracted {len(roles_data)} roles with descriptions\n")

        # Display the extracted data
        for i, role in enumerate(roles_data, 1):
            print(f"\n{i}. {role['name']}")
            print(f"   Description: {role['description'][:100]}..." if len(role['description']) > 100 else f"   Description: {role['description']}")

        return roles_data

    except FileNotFoundError:
        print(f"ERROR: Excel file not found at: {EXCEL_PATH}")
        return None
    except Exception as e:
        print(f"ERROR reading Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None


def update_role_descriptions(roles_data):
    """Update the role_cluster table with descriptions from Excel"""
    print("\n\n=== Updating Database ===\n")

    app = create_app()
    with app.app_context():
        updated_count = 0
        not_found_count = 0

        for role_data in roles_data:
            role_name = role_data['name']
            description = role_data['description']

            # Try to find the role in the database
            role = RoleCluster.query.filter_by(role_cluster_name=role_name).first()

            if role:
                # Update the description
                old_desc = role.role_cluster_description
                role.role_cluster_description = description
                updated_count += 1
                print(f"[OK] Updated: {role_name}")
                print(f"     Old: {old_desc[:80] if old_desc else 'None'}...")
                print(f"     New: {description[:80]}...\n")
            else:
                not_found_count += 1
                print(f"[WARNING] Role not found in database: {role_name}")

        # Commit the changes
        try:
            db.session.commit()
            print(f"\n=== Summary ===")
            print(f"Updated: {updated_count} roles")
            print(f"Not found: {not_found_count} roles")
            print(f"\nChanges committed to database successfully!\n")
        except Exception as e:
            db.session.rollback()
            print(f"\nERROR committing to database: {e}")
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    # Check for command-line argument to auto-confirm
    import sys
    auto_confirm = len(sys.argv) > 1 and sys.argv[1] == '--yes'

    # Extract role descriptions from Excel
    roles_data = extract_role_descriptions()

    if roles_data:
        print("\n" + "="*60)
        print("Ready to update database with these descriptions")
        print("="*60)

        if auto_confirm:
            print("\nAuto-confirming update (--yes flag provided)")
            update_role_descriptions(roles_data)
        else:
            # Ask for confirmation
            response = input("\nProceed with database update? (yes/no): ").strip().lower()

            if response in ['yes', 'y']:
                update_role_descriptions(roles_data)
            else:
                print("\nDatabase update cancelled by user.")
    else:
        print("\nNo data extracted from Excel. Aborting.")
